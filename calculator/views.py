from django.shortcuts import render, redirect
from django.http import FileResponse, HttpResponse
from django.conf import settings
import os
import openpyxl

from .models import Consumer
from .forms import ConsumerForm



# TODO: Your list view should do the following tasks
"""
-> Recover all consumers from the database
-> Get the discount value for each consumer
-> Calculate the economy
-> Send the data to the template that will be rendered
"""


def list_view(request):
    list_consumers = Consumer.objects.all()
    count_consumers = Consumer.objects.count()

    context = {
        'list_consumers': list_consumers,
        'count_consumers': count_consumers,
    }

    return render(request, 'calculator/list.html', context=context)
    # Create the first view here.


# TODO: Your create view should do the following tasks
"""Create a view to perform inclusion of consumers. The view should do:
-> Receive a POST request with the data to register
-> If the data is valid (validate document), create and save a new Consumer object associated with the right discount rule object
-> Redirect to the template that list all consumers

Your view must be associated with an url and a template different from the first one. A link to
this page must be provided in the main page.
"""


def register_view(request):
    list_consumers = Consumer.objects.all()
    count_consumers = Consumer.objects.count()

    def redirect_into_register_view(): 
        print("clickou")

    context = {
        'list_consumers': list_consumers,
        'count_consumers': count_consumers,
        'redirect_into_register_view': redirect_into_register_view
    }

    return render(request, 'calculator/register.html', context=context)

def register_view(request):
    if request.method == 'POST':
        form = ConsumerForm(request.POST)
        if form.is_valid():
            form.save()

            name = form.cleaned_data['name']
            document = form.cleaned_data['document']
            zip_code = form.cleaned_data['zip_code']
            city = form.cleaned_data['city']
            state = form.cleaned_data['state']
            consumption = form.cleaned_data['consumption']
            distributor_tax = form.cleaned_data['distributor_tax']

            workbook = openpyxl.load_workbook('consumers.xlsx')
            sheet = workbook.active

            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=name)
            sheet.cell(row=next_row, column=2, value=document)
            sheet.cell(row=next_row, column=3, value=zip_code)
            sheet.cell(row=next_row, column=4, value=city)
            sheet.cell(row=next_row, column=5, value=state)
            sheet.cell(row=next_row, column=6, value=consumption)
            sheet.cell(row=next_row, column=7, value=distributor_tax)

            workbook.save('consumers.xlsx')

            return redirect('list')
    else:
        form = ConsumerForm()
    return render(request, 'calculator/register.html', {'form': form})

def download_excel(request):
    excel_file_path = os.path.join(settings.BASE_DIR, 'consumers.xlsx')

    if os.path.exists(excel_file_path):
        with open(excel_file_path, 'rb') as excel_file:
            response = FileResponse(excel_file, as_attachment=True, filename='consumers.xlsx')
            return response
    else:
        return HttpResponse('Arquivo não encontrado', status=404)
